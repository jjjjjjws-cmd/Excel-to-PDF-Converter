#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel to PDF 변환기 v3.2.3
시트별 열+폴더 선택 - 한 줄 UI
"""

import os
import time
import base64
import threading
import hashlib
import openpyxl
from tkinter import *
from tkinter import ttk, messagebox, filedialog
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

class ExcelToPDFApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel to PDF 변환기 v3.2.3")
        self.root.geometry("800x750")
        
        self.excel_path = None
        self.wb = None
        self.is_running = False
        self.mode = IntVar(value=0)
        self.threads = []
        self.drivers = []
        
        self.sheet_rows = {}
        
        self.setup_ui()
        
    def setup_ui(self):
        title_frame = Frame(self.root, bg="#2196F3", height=60)
        title_frame.pack(fill=X)
        Label(title_frame, text="Excel to PDF 변환기 v3.2.3", 
              font=("맑은 고딕", 16, "bold"), bg="#2196F3", fg="white").pack(pady=15)
        
        main_frame = Frame(self.root, padx=20, pady=10)
        main_frame.pack(fill=BOTH, expand=True)
        
        mode_frame = Frame(main_frame)
        mode_frame.pack(fill=X, pady=(0,15))
        
        Label(mode_frame, text="모드 선택:", font=("맑은 고딕", 10, "bold")).pack(side=LEFT, padx=(0,20))
        Radiobutton(mode_frame, text="간단 모드", variable=self.mode, value=0, 
                   command=self.switch_mode, font=("맑은 고딕", 9)).pack(side=LEFT, padx=5)
        Radiobutton(mode_frame, text="고급 모드", variable=self.mode, value=1, 
                   command=self.switch_mode, font=("맑은 고딕", 9)).pack(side=LEFT, padx=5)
        
        self.content_frame = Frame(main_frame)
        self.content_frame.pack(fill=BOTH, expand=True)
        
        self.setup_simple_mode()
        
    def setup_simple_mode(self):
        for widget in self.content_frame.winfo_children():
            widget.destroy()
        
        Label(self.content_frame, text="1. 엑셀 파일", font=("맑은 고딕", 10, "bold")).pack(anchor="w", pady=(0,5))
        file_frame = Frame(self.content_frame)
        file_frame.pack(fill=X, pady=(0,15))
        
        self.simple_file_label = Label(file_frame, text="파일을 선택하세요", bg="#f0f0f0", 
                                       anchor="w", padx=10, pady=8, relief="solid", borderwidth=1)
        self.simple_file_label.pack(side=LEFT, fill=X, expand=True, padx=(0,10))
        
        Button(file_frame, text="선택", command=self.select_file_simple, 
               bg="#4CAF50", fg="white", padx=15, pady=5).pack(side=RIGHT)
        
        Label(self.content_frame, text="2. 링크가 있는 열", font=("맑은 고딕", 10, "bold")).pack(anchor="w", pady=(0,5))
        col_frame = Frame(self.content_frame)
        col_frame.pack(fill=X, pady=(0,15))
        
        Label(col_frame, text="열 (예: H):").pack(side=LEFT)
        self.simple_column_entry = Entry(col_frame, width=10)
        self.simple_column_entry.pack(side=LEFT, padx=(10,0))
        self.simple_column_entry.insert(0, "H")
        
        Label(self.content_frame, text="3. 시트 선택", font=("맑은 고딕", 10, "bold")).pack(anchor="w", pady=(0,5))
        
        sheet_frame = Frame(self.content_frame)
        sheet_frame.pack(fill=BOTH, expand=True, pady=(0,10))
        
        scrollbar = Scrollbar(sheet_frame, orient=VERTICAL)
        self.simple_sheet_listbox = Listbox(sheet_frame, selectmode=MULTIPLE, 
                                            yscrollcommand=scrollbar.set, height=4)
        scrollbar.config(command=self.simple_sheet_listbox.yview)
        scrollbar.pack(side=RIGHT, fill=Y)
        self.simple_sheet_listbox.pack(side=LEFT, fill=BOTH, expand=True)
        
        Button(self.content_frame, text="전체 선택", 
               command=lambda: self.simple_sheet_listbox.select_set(0, END)).pack(pady=(0,15))
        
        Label(self.content_frame, text="4. 저장 폴더", font=("맑은 고딕", 10, "bold")).pack(anchor="w", pady=(0,5))
        folder_frame = Frame(self.content_frame)
        folder_frame.pack(fill=X, pady=(0,20))
        
        self.simple_folder_label = Label(folder_frame, text="폴더를 선택하세요", bg="#f0f0f0",
                                         anchor="w", padx=10, pady=8, relief="solid", borderwidth=1)
        self.simple_folder_label.pack(side=LEFT, fill=X, expand=True, padx=(0,10))
        
        Button(folder_frame, text="선택", command=self.select_folder_simple,
               bg="#2196F3", fg="white", padx=15, pady=5).pack(side=RIGHT)
        
        self.simple_start_button = Button(self.content_frame, text="변환 시작", 
                                          command=self.start_simple,
                                          bg="#FF5722", fg="white", font=("맑은 고딕", 12, "bold"),
                                          padx=30, pady=10)
        self.simple_start_button.pack(pady=10)
        
        Label(self.content_frame, text="진행 상황", font=("맑은 고딕", 10, "bold")).pack(anchor="w", pady=(10,5))
        self.simple_progress = ttk.Progressbar(self.content_frame, mode='determinate')
        self.simple_progress.pack(fill=X, pady=(0,5))
        
        self.simple_status_label = Label(self.content_frame, text="대기 중...", fg="#666")
        self.simple_status_label.pack(anchor="w")
        
    def setup_multi_mode(self):
        for widget in self.content_frame.winfo_children():
            widget.destroy()
        
        Label(self.content_frame, text="1. 엑셀 파일", font=("맑은 고딕", 10, "bold")).pack(anchor="w", pady=(0,5))
        file_frame = Frame(self.content_frame)
        file_frame.pack(fill=X, pady=(0,15))
        
        self.multi_file_label = Label(file_frame, text="파일을 선택하세요", bg="#f0f0f0", 
                                      anchor="w", padx=10, pady=8, relief="solid", borderwidth=1)
        self.multi_file_label.pack(side=LEFT, fill=X, expand=True, padx=(0,10))
        
        Button(file_frame, text="선택", command=self.select_file_multi, 
               bg="#4CAF50", fg="white", padx=15, pady=5).pack(side=RIGHT)
        
        Label(self.content_frame, text="2. 시트 및 열 설정", font=("맑은 고딕", 10, "bold")).pack(anchor="w", pady=(0,5))
        
        canvas_frame = Frame(self.content_frame, relief="solid", borderwidth=1)
        canvas_frame.pack(fill=BOTH, expand=True, pady=(0,15))
        
        canvas = Canvas(canvas_frame, height=300)
        scrollbar = Scrollbar(canvas_frame, orient=VERTICAL, command=canvas.yview)
        self.multi_sheets_frame = Frame(canvas)
        
        self.multi_sheets_frame.bind("<Configure>", 
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        
        canvas.create_window((0, 0), window=self.multi_sheets_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar.pack(side=RIGHT, fill=Y)
        
        self.multi_start_button = Button(self.content_frame, text="변환 시작", 
                                         command=self.start_multi,
                                         bg="#FF5722", fg="white", font=("맑은 고딕", 12, "bold"),
                                         padx=30, pady=10)
        self.multi_start_button.pack(pady=10)
        
        Label(self.content_frame, text="진행 상황", font=("맑은 고딕", 10, "bold")).pack(anchor="w", pady=(10,5))
        
        progress_canvas_frame = Frame(self.content_frame, relief="solid", borderwidth=1)
        progress_canvas_frame.pack(fill=BOTH, expand=True)
        
        progress_canvas = Canvas(progress_canvas_frame, height=150)
        progress_scrollbar = Scrollbar(progress_canvas_frame, orient=VERTICAL, command=progress_canvas.yview)
        self.multi_progress_frame = Frame(progress_canvas)
        
        self.multi_progress_frame.bind("<Configure>",
            lambda e: progress_canvas.configure(scrollregion=progress_canvas.bbox("all")))
        
        progress_canvas.create_window((0, 0), window=self.multi_progress_frame, anchor="nw")
        progress_canvas.configure(yscrollcommand=progress_scrollbar.set)
        
        progress_canvas.pack(side=LEFT, fill=BOTH, expand=True)
        progress_scrollbar.pack(side=RIGHT, fill=Y)
    
    def switch_mode(self):
        if self.is_running:
            messagebox.showwarning("경고", "작업 중에는 모드를 변경할 수 없습니다.")
            return
        
        if self.mode.get() == 0:
            self.setup_simple_mode()
        else:
            self.setup_multi_mode()
    
    def select_file_simple(self):
        filename = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if filename:
            self.excel_path = filename
            self.simple_file_label.config(text=os.path.basename(filename))
            self.load_sheets_simple()
    
    def load_sheets_simple(self):
        try:
            if self.wb:
                self.wb.close()
            
            self.wb = openpyxl.load_workbook(self.excel_path, data_only=True)
            self.simple_sheet_listbox.delete(0, END)
            
            for idx, name in enumerate(self.wb.sheetnames, 1):
                ws = self.wb[name]
                self.simple_sheet_listbox.insert(END, str(idx) + ". " + name + " (" + str(ws.max_row) + "개)")
                
        except Exception as e:
            messagebox.showerror("오류", "파일 로드 실패:\n" + str(e))
            self.wb = None
    
    def select_folder_simple(self):
        folder = filedialog.askdirectory(title="저장 폴더")
        if folder:
            self.simple_folder = folder
            display = folder if len(folder) < 45 else "..." + folder[-42:]
            self.simple_folder_label.config(text=display)
    
    def select_file_multi(self):
        filename = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        
        if filename:
            self.excel_path = filename
            self.multi_file_label.config(text=os.path.basename(filename))
            self.load_sheets_multi()
    
    def load_sheets_multi(self):
        try:
            if self.wb:
                self.wb.close()
            
            self.wb = openpyxl.load_workbook(self.excel_path, data_only=True)
            
            for widget in self.multi_sheets_frame.winfo_children():
                widget.destroy()
            
            self.sheet_rows = {}
            
            for sheet_name in self.wb.sheetnames:
                ws = self.wb[sheet_name]
                
                row_frame = Frame(self.multi_sheets_frame, pady=5)
                row_frame.pack(fill=X, padx=5, pady=2)
                
                var = IntVar(value=1)
                cb = Checkbutton(row_frame, text=sheet_name + " (" + str(ws.max_row) + "개)", 
                                variable=var, font=("맑은 고딕", 9))
                cb.pack(side=LEFT)
                
                Label(row_frame, text=" - 열:", font=("맑은 고딕", 9)).pack(side=LEFT, padx=(5,5))
                
                col_container = Frame(row_frame)
                col_container.pack(side=LEFT, fill=X, expand=True)
                
                self.add_column_entry(col_container, "H")
                
                Button(row_frame, text="+", command=lambda cc=col_container: self.add_column_entry(cc, ""),
                       bg="#4CAF50", fg="white", padx=8, pady=2, font=("맑은 고딕", 9, "bold")).pack(side=LEFT, padx=2)
                
                self.sheet_rows[sheet_name] = {
                    'var': var,
                    'col_container': col_container,
                    'row_count': ws.max_row
                }
                
        except Exception as e:
            messagebox.showerror("오류", "파일 로드 실패:\n" + str(e))
            self.wb = None
    
    def add_column_entry(self, container, default_val):
        entry_box = Frame(container)
        entry_box.pack(side=LEFT, padx=2)
        
        col_entry = Entry(entry_box, width=3, font=("맑은 고딕", 9))
        col_entry.pack(side=LEFT)
        if default_val:
            col_entry.insert(0, default_val)
        
        folder_info = {"path": None}
        
        folder_btn = Button(entry_box, text="📁", 
                           command=lambda: self.pick_folder_for_col(col_entry, folder_info, folder_btn),
                           bg="#2196F3", fg="white", padx=3, pady=0, font=("맑은 고딕", 8))
        folder_btn.pack(side=LEFT, padx=2)
        
        Button(entry_box, text="×", command=lambda: self.remove_column_entry(container, entry_box),
               fg="red", padx=5, pady=0, font=("맑은 고딕", 8)).pack(side=LEFT, padx=2)
        
        entry_box.folder_info = folder_info
    
    def remove_column_entry(self, container, entry_box):
        if len(container.winfo_children()) > 1:
            entry_box.destroy()
        else:
            messagebox.showwarning("경고", "최소 1개의 열은 필요합니다.")
    
    def pick_folder_for_col(self, entry, folder_info, button):
        col = entry.get().strip().upper()
        if not col:
            messagebox.showwarning("경고", "먼저 열을 입력하세요")
            return
        
        folder = filedialog.askdirectory(title=col + "열 저장 폴더 선택")
        if folder:
            folder_info["path"] = folder
            button.config(bg="#4CAF50")
    
    def start_simple(self):
        if not self.excel_path or not self.wb:
            messagebox.showwarning("경고", "파일을 선택하세요")
            return
        if not self.simple_column_entry.get().strip():
            messagebox.showwarning("경고", "열을 입력하세요")
            return
        if not self.simple_sheet_listbox.curselection():
            messagebox.showwarning("경고", "시트를 선택하세요")
            return
        if not hasattr(self, 'simple_folder'):
            messagebox.showwarning("경고", "폴더를 선택하세요")
            return
        
        if self.is_running:
            return
        
        self.is_running = True
        self.simple_start_button.config(state="disabled")
        
        thread = threading.Thread(target=self.process_simple, daemon=True)
        thread.start()
    
    def process_simple(self):
        driver = None
        wb_local = None
        try:
            column = self.simple_column_entry.get().strip().upper()
            col_idx = self.col_to_idx(column)
            selected = self.simple_sheet_listbox.curselection()
            
            wb_local = openpyxl.load_workbook(self.excel_path, data_only=True)
            
            options = Options()
            options.add_argument('--incognito')
            options.add_argument('--disable-blink-features=AutomationControlled')
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            
            driver = webdriver.Chrome(options=options)
            self.drivers.append(driver)
            
            all_links = []
            for idx in selected:
                sheet_name = wb_local.sheetnames[idx]
                ws = wb_local[sheet_name]
                
                for row in range(4, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value and str(cell.value).startswith('http'):
                        store = ws.cell(row=row, column=4).value or ""
                        all_links.append({'url': cell.value, 'store': store})
            
            total = len(all_links)
            
            if total == 0:
                messagebox.showwarning("경고", "링크가 없습니다")
                return
            
            for i, link in enumerate(all_links, 1):
                if not self.is_running:
                    break
                
                progress = int(i / total * 100)
                self.update_simple_progress(progress, str(i) + "/" + str(total) + " - " + link['store'])
                
                self.save_pdf(driver, link['url'], self.simple_folder)
                time.sleep(0.5)
            
            self.update_simple_progress(100, "완료!")
            messagebox.showinfo("완료", "총 " + str(total) + "개 처리 완료")
            
        except Exception as e:
            messagebox.showerror("오류", "작업 실패:\n" + str(e))
        
        finally:
            if wb_local:
                try:
                    wb_local.close()
                except:
                    pass
            
            if driver:
                try:
                    driver.quit()
                except:
                    pass
                if driver in self.drivers:
                    self.drivers.remove(driver)
            
            self.simple_start_button.config(state="normal")
            self.is_running = False
    
    def start_multi(self):
        if not self.excel_path or not self.wb:
            messagebox.showwarning("경고", "파일을 선택하세요")
            return
        
        tasks = []
        for sheet_name, sheet_info in self.sheet_rows.items():
            if sheet_info['var'].get() == 0:
                continue
            
            for entry_box in sheet_info['col_container'].winfo_children():
                widgets = entry_box.winfo_children()
                if len(widgets) < 1:
                    continue
                
                col_entry = widgets[0]
                col = col_entry.get().strip().upper()
                
                if not col:
                    continue
                
                if not hasattr(entry_box, 'folder_info') or not entry_box.folder_info.get("path"):
                    messagebox.showwarning("경고", sheet_name + " - " + col + "열의 폴더를 선택하세요")
                    return
                
                folder = entry_box.folder_info["path"]
                
                task_data = {}
                task_data['sheet'] = sheet_name
                task_data['column'] = col
                task_data['folder'] = folder
                task_data['label'] = sheet_name[:10] + "-" + col
                tasks.append(task_data)
        
        if not tasks:
            messagebox.showwarning("경고", "처리할 작업이 없습니다")
            return
        
        if self.is_running:
            return
        
        self.is_running = True
        self.multi_start_button.config(state="disabled")
        
        for widget in self.multi_progress_frame.winfo_children():
            widget.destroy()
        
        self.progress_bars = {}
        for task in tasks:
            frame = Frame(self.multi_progress_frame)
            frame.pack(fill=X, pady=2)
            
            Label(frame, text="[" + task['label'] + "]", width=18, anchor="w").pack(side=LEFT)
            
            progress = ttk.Progressbar(frame, mode='determinate')
            progress.pack(side=LEFT, fill=X, expand=True, padx=5)
            
            status = Label(frame, text="0%", width=10)
            status.pack(side=RIGHT)
            
            bar_data = {}
            bar_data['bar'] = progress
            bar_data['label'] = status
            self.progress_bars[task['label']] = bar_data
        
        self.threads = []
        for task in tasks:
            thread = threading.Thread(target=self.process_multi, args=(task,), daemon=True)
            thread.start()
            self.threads.append(thread)
    
    def process_multi(self, task):
        driver = None
        wb_local = None
        try:
            wb_local = openpyxl.load_workbook(self.excel_path, data_only=True)
            
            options = Options()
            options.add_argument('--incognito')
            options.add_argument('--disable-blink-features=AutomationControlled')
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            
            driver = webdriver.Chrome(options=options)
            self.drivers.append(driver)
            
            ws = wb_local[task['sheet']]
            col_idx = self.col_to_idx(task['column'])
            
            links = []
            for row in range(4, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if cell.value and str(cell.value).startswith('http'):
                    store = ws.cell(row=row, column=4).value or ""
                    links.append({'url': cell.value, 'store': store})
            
            total = len(links)
            
            if total == 0:
                self.update_multi_progress(task['label'], 0, "링크없음")
                return
            
            for i, link in enumerate(links, 1):
                if not self.is_running:
                    break
                
                progress = int(i / total * 100)
                self.update_multi_progress(task['label'], progress, str(i) + "/" + str(total))
                
                self.save_pdf(driver, link['url'], task['folder'])
                time.sleep(0.3)
            
            self.update_multi_progress(task['label'], 100, "✓")
            
        except Exception as e:
            self.update_multi_progress(task['label'], 0, "✗")
        
        finally:
            if wb_local:
                try:
                    wb_local.close()
                except:
                    pass
            
            if driver:
                try:
                    driver.quit()
                except:
                    pass
                if driver in self.drivers:
                    self.drivers.remove(driver)
            
            if all(not t.is_alive() for t in self.threads if t != threading.current_thread()):
                self.multi_start_button.config(state="normal")
                self.is_running = False
                messagebox.showinfo("완료", "모든 작업 완료!")
    
    def save_pdf(self, driver, url, folder):
        try:
            driver.get(url)
            time.sleep(2)
            
            try:
                title = driver.title
                title = "".join(c for c in title if c.isalnum() or c in (' ', '_', '-')).strip()
                if not title:
                    title = "document"
            except:
                title = hashlib.md5(url.encode()).hexdigest()[:8]
            
            filename = title + ".pdf"
            filepath = os.path.join(folder, filename)
            
            counter = 1
            while os.path.exists(filepath):
                filename = title + "_" + str(counter) + ".pdf"
                filepath = os.path.join(folder, filename)
                counter += 1
            
            result = driver.execute_cdp_cmd('Page.printToPDF', {
                'landscape': False,
                'displayHeaderFooter': False,
                'printBackground': True,
                'preferCSSPageSize': True,
            })
            
            pdf_data = base64.b64decode(result['data'])
            
            with open(filepath, 'wb') as f:
                f.write(pdf_data)
            
            return True
            
        except Exception as e:
            print("PDF 저장 실패: " + str(e))
            return False
    
    def update_simple_progress(self, value, message):
        self.simple_progress['value'] = value
        self.simple_status_label.config(text=message)
        self.root.update()
    
    def update_multi_progress(self, label, value, message):
        if label in self.progress_bars:
            self.progress_bars[label]['bar']['value'] = value
            self.progress_bars[label]['label'].config(text=message)
            self.root.update()
    
    def col_to_idx(self, col):
        result = 0
        for c in col:
            result = result * 26 + (ord(c) - ord('A') + 1)
        return result

def main():
    root = Tk()
    app = ExcelToPDFApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
